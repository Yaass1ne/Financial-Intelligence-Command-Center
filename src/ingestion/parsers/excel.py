"""
Excel Budget Parser

Parses budget Excel files with support for merged cells, multiple sheets,
and various amount formats.

Examples:
    >>> from pathlib import Path
    >>> budget = parse_budget_excel(Path("budget_2024_Q1.xlsx"))
    >>> print(budget['department'].unique())
"""

import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
from pathlib import Path
from typing import Dict, Any, List, Optional, Tuple
import logging
import re

from src.ingestion.extractors.amounts import parse_amount
from src.ingestion.extractors.dates import parse_date

logger = logging.getLogger(__name__)


class ExcelParseError(Exception):
    """Exception raised when Excel parsing fails."""
    pass


def parse_budget_excel(excel_path: Path, sheet_name: Optional[str] = None) -> pd.DataFrame:
    """
    Parse budget Excel file into structured dataframe.

    Handles:
    - Merged cells
    - Multiple sheets
    - Various amount formats (1,250.50 vs 1.250,50)
    - Header detection
    - Empty rows

    Args:
        excel_path: Path to Excel file
        sheet_name: Specific sheet to parse (default: first sheet or "Budget")

    Returns:
        DataFrame with columns: department, category, budget, actual, variance, period

    Raises:
        ExcelParseError: If file cannot be parsed

    Examples:
        >>> budget_df = parse_budget_excel(Path("budget_2024.xlsx"))
        >>> print(budget_df.head())
    """
    if not excel_path.exists():
        raise ExcelParseError(f"File not found: {excel_path}")

    try:
        # Load workbook with openpyxl to handle merged cells
        wb = openpyxl.load_workbook(excel_path, data_only=True)

        # Determine which sheet to parse
        if sheet_name:
            if sheet_name not in wb.sheetnames:
                raise ExcelParseError(f"Sheet '{sheet_name}' not found in {excel_path}")
            ws = wb[sheet_name]
        else:
            # Try to find "Budget" sheet, otherwise use first sheet
            if "Budget" in wb.sheetnames:
                ws = wb["Budget"]
            elif "budget" in [s.lower() for s in wb.sheetnames]:
                sheet_name = [s for s in wb.sheetnames if s.lower() == "budget"][0]
                ws = wb[sheet_name]
            else:
                ws = wb.active

        logger.info(f"Parsing sheet '{ws.title}' from {excel_path.name}")

        # Unmerge cells and fill with top-left value
        merged_ranges = list(ws.merged_cells.ranges)
        for merged_range in merged_ranges:
            # Get top-left cell value
            top_left_cell = ws.cell(merged_range.min_row, merged_range.min_col)
            value = top_left_cell.value

            # Unmerge
            ws.unmerge_cells(str(merged_range))

            # Fill all cells in the range with the value
            for row in range(merged_range.min_row, merged_range.max_row + 1):
                for col in range(merged_range.min_col, merged_range.max_col + 1):
                    ws.cell(row, col).value = value

        # Convert to pandas DataFrame
        data = []
        for row in ws.iter_rows(values_only=True):
            data.append(row)

        df = pd.DataFrame(data)

        # Find header row
        header_row = _find_header_row(df)
        if header_row is None:
            logger.warning(f"Could not find header row in {excel_path.name}, using first row")
            header_row = 0

        # Set headers
        df.columns = df.iloc[header_row]
        df = df.iloc[header_row + 1:].reset_index(drop=True)

        # Normalize column names
        df.columns = [_normalize_column_name(col) for col in df.columns]

        # Remove empty rows
        df = df.dropna(how='all')

        # Detect and map columns to standard schema
        column_mapping = _detect_column_mapping(df.columns)

        # Rename columns to standard names
        df = df.rename(columns=column_mapping)

        # Parse and normalize amounts
        for col in ['budget', 'actual', 'forecast', 'variance']:
            if col in df.columns:
                df[col] = df[col].apply(_normalize_amount)

        # Parse dates if present
        for col in ['period', 'date', 'month']:
            if col in df.columns:
                df[col] = df[col].apply(_normalize_date)

        # Calculate variance if missing
        if 'variance' not in df.columns and 'budget' in df.columns and 'actual' in df.columns:
            df['variance'] = df['actual'] - df['budget']
            df['variance_percent'] = (df['variance'] / df['budget'] * 100).round(2)

        # Add metadata
        df['source_file'] = str(excel_path)
        df['sheet_name'] = ws.title

        logger.info(f"Parsed {len(df)} budget rows from {excel_path.name}")
        return df

    except Exception as e:
        raise ExcelParseError(f"Error parsing {excel_path}: {e}")


def parse_multi_sheet_budget(excel_path: Path) -> Dict[str, pd.DataFrame]:
    """
    Parse all sheets from budget Excel file.

    Args:
        excel_path: Path to Excel file

    Returns:
        Dictionary mapping sheet name to DataFrame

    Examples:
        >>> sheets = parse_multi_sheet_budget(Path("budget_2024.xlsx"))
        >>> for sheet_name, df in sheets.items():
        ...     print(f"{sheet_name}: {len(df)} rows")
    """
    if not excel_path.exists():
        raise ExcelParseError(f"File not found: {excel_path}")

    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        results = {}

        for sheet_name in wb.sheetnames:
            try:
                df = parse_budget_excel(excel_path, sheet_name=sheet_name)
                results[sheet_name] = df
            except Exception as e:
                logger.warning(f"Could not parse sheet '{sheet_name}': {e}")
                continue

        return results

    except Exception as e:
        raise ExcelParseError(f"Error parsing {excel_path}: {e}")


def detect_table_structure(excel_path: Path, sheet_name: Optional[str] = None) -> Dict[str, Any]:
    """
    Automatically detect table structure in Excel file.

    Args:
        excel_path: Path to Excel file
        sheet_name: Specific sheet to analyze

    Returns:
        Dictionary with table metadata (header_row, num_columns, num_rows, etc.)
    """
    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        ws = wb[sheet_name] if sheet_name else wb.active

        # Find dimensions
        max_row = ws.max_row
        max_col = ws.max_column

        # Find header row (row with most non-empty cells)
        header_row = 0
        max_filled = 0

        for row_idx in range(1, min(10, max_row + 1)):  # Check first 10 rows
            filled = sum(1 for cell in ws[row_idx] if cell.value is not None)
            if filled > max_filled:
                max_filled = filled
                header_row = row_idx

        # Count data rows (after header)
        data_rows = 0
        for row_idx in range(header_row + 1, max_row + 1):
            if any(cell.value is not None for cell in ws[row_idx]):
                data_rows += 1

        return {
            'header_row': header_row,
            'num_columns': max_col,
            'num_rows': data_rows,
            'total_rows': max_row,
            'sheet_name': ws.title,
            'has_merged_cells': len(list(ws.merged_cells.ranges)) > 0
        }

    except Exception as e:
        logger.error(f"Error detecting table structure in {excel_path}: {e}")
        return {}


# Helper functions

def _find_header_row(df: pd.DataFrame, max_rows_to_check: int = 10) -> Optional[int]:
    """
    Find the row that contains column headers.

    Looks for rows with:
    - Mostly text values
    - Keywords like "department", "budget", "actual", etc.

    Args:
        df: DataFrame to analyze
        max_rows_to_check: Maximum number of rows to check

    Returns:
        Index of header row, or None if not found
    """
    header_keywords = [
        'department', 'departement', 'service',
        'budget', 'actual', 'reel', 'réel',
        'category', 'categorie', 'catégorie',
        'amount', 'montant', 'total',
        'period', 'periode', 'période', 'date'
    ]

    for idx in range(min(max_rows_to_check, len(df))):
        row = df.iloc[idx]

        # Convert to lowercase strings
        row_str = [str(val).lower() if val is not None else '' for val in row]

        # Count matches with header keywords
        matches = sum(1 for val in row_str if any(keyword in val for keyword in header_keywords))

        # If more than 2 keywords match, likely a header row
        if matches >= 2:
            return idx

    return None


def _normalize_column_name(col: Any) -> str:
    """
    Normalize column name to lowercase with underscores.

    Args:
        col: Column name (can be string, number, or None)

    Returns:
        Normalized column name
    """
    if col is None:
        return 'unnamed'

    col_str = str(col).strip().lower()

    # Remove special characters
    col_str = re.sub(r'[^\w\s]', '', col_str)

    # Replace spaces with underscores
    col_str = re.sub(r'\s+', '_', col_str)

    return col_str


def _detect_column_mapping(columns: List[str]) -> Dict[str, str]:
    """
    Map detected columns to standard schema.

    Args:
        columns: List of column names from Excel

    Returns:
        Dictionary mapping original names to standard names
    """
    mapping = {}

    # Define patterns for each standard column
    patterns = {
        'department': ['department', 'departement', 'service', 'dept'],
        'category': ['category', 'categorie', 'catégorie', 'type', 'poste'],
        'budget': ['budget', 'budgeted', 'budgete', 'budgété', 'planned', 'prévu', 'prevu'],
        'actual': ['actual', 'reel', 'réel', 'spent', 'dépensé', 'depense'],
        'forecast': ['forecast', 'prévision', 'prevision', 'estimated', 'estimé', 'estime'],
        'variance': ['variance', 'écart', 'ecart', 'difference', 'différence'],
        'period': ['period', 'période', 'periode', 'month', 'mois', 'date', 'quarter', 'trimestre'],
        'notes': ['notes', 'comments', 'commentaires', 'remarks']
    }

    for col in columns:
        col_lower = col.lower() if isinstance(col, str) else str(col).lower()

        # Check each pattern
        for standard_name, keywords in patterns.items():
            if any(keyword in col_lower for keyword in keywords):
                mapping[col] = standard_name
                break

    return mapping


def _normalize_amount(value: Any) -> Optional[float]:
    """
    Normalize amount value to float.

    Args:
        value: Amount value (can be string, number, or None)

    Returns:
        Normalized amount as float, or None
    """
    if value is None or value == '':
        return None

    if isinstance(value, (int, float)):
        return float(value)

    if isinstance(value, str):
        # Use the amount parser
        return parse_amount(value)

    return None


def _normalize_date(value: Any) -> Optional[str]:
    """
    Normalize date value to ISO format string.

    Args:
        value: Date value (can be string, datetime, or None)

    Returns:
        ISO format date string (YYYY-MM-DD), or None
    """
    if value is None or value == '':
        return None

    # If already a pandas Timestamp or datetime
    if isinstance(value, (pd.Timestamp, pd.datetime)):
        return value.strftime('%Y-%m-%d')

    # If string, parse it
    if isinstance(value, str):
        parsed = parse_date(value)
        if parsed:
            return parsed.strftime('%Y-%m-%d')

    return None


def extract_budget_summary(df: pd.DataFrame) -> Dict[str, Any]:
    """
    Extract summary statistics from budget DataFrame.

    Args:
        df: Budget DataFrame

    Returns:
        Dictionary with summary statistics

    Examples:
        >>> summary = extract_budget_summary(budget_df)
        >>> print(f"Total budget: {summary['total_budget']}")
    """
    summary = {}

    # Total amounts
    if 'budget' in df.columns:
        summary['total_budget'] = df['budget'].sum()

    if 'actual' in df.columns:
        summary['total_actual'] = df['actual'].sum()

    if 'variance' in df.columns:
        summary['total_variance'] = df['variance'].sum()
    elif 'budget' in df.columns and 'actual' in df.columns:
        summary['total_variance'] = df['actual'].sum() - df['budget'].sum()

    # Calculate variance percentage
    if 'total_budget' in summary and 'total_variance' in summary and summary['total_budget'] != 0:
        summary['total_variance_percent'] = (summary['total_variance'] / summary['total_budget']) * 100

    # Department breakdown
    if 'department' in df.columns and 'budget' in df.columns:
        summary['by_department'] = df.groupby('department')['budget'].sum().to_dict()

    # Category breakdown
    if 'category' in df.columns and 'budget' in df.columns:
        summary['by_category'] = df.groupby('category')['budget'].sum().to_dict()

    # Number of line items
    summary['num_items'] = len(df)

    # Departments with overruns
    if 'department' in df.columns and 'variance' in df.columns:
        overruns = df[df['variance'] > 0].groupby('department')['variance'].sum()
        summary['overruns_by_department'] = overruns.to_dict()

    return summary


def validate_budget_data(df: pd.DataFrame) -> List[Dict[str, Any]]:
    """
    Validate budget data and return list of issues.

    Args:
        df: Budget DataFrame

    Returns:
        List of validation issues

    Examples:
        >>> issues = validate_budget_data(budget_df)
        >>> for issue in issues:
        ...     print(f"{issue['severity']}: {issue['message']}")
    """
    issues = []

    # Check for required columns
    required_columns = ['department', 'budget']
    for col in required_columns:
        if col not in df.columns:
            issues.append({
                'severity': 'ERROR',
                'column': col,
                'message': f"Required column '{col}' is missing"
            })

    # Check for negative budgets
    if 'budget' in df.columns:
        negative_budgets = df[df['budget'] < 0]
        if not negative_budgets.empty:
            issues.append({
                'severity': 'WARNING',
                'column': 'budget',
                'count': len(negative_budgets),
                'message': f"Found {len(negative_budgets)} rows with negative budget values"
            })

    # Check for null values in important columns
    for col in ['department', 'budget', 'actual']:
        if col in df.columns:
            null_count = df[col].isnull().sum()
            if null_count > 0:
                issues.append({
                    'severity': 'WARNING',
                    'column': col,
                    'count': null_count,
                    'message': f"Found {null_count} null values in column '{col}'"
                })

    # Check for duplicate departments (same period)
    if 'department' in df.columns and 'period' in df.columns:
        duplicates = df.groupby(['department', 'period']).size()
        duplicates = duplicates[duplicates > 1]
        if not duplicates.empty:
            issues.append({
                'severity': 'WARNING',
                'column': 'department',
                'count': len(duplicates),
                'message': f"Found {len(duplicates)} duplicate department-period combinations"
            })

    return issues
